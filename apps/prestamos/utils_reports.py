import csv
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Count
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from catalogo.models import Libro
from usuarios.models import Lector
from .models import Prestamo

def exportar_inventario_excel():
    """Genera un reporte Excel completo del inventario de libros y estadísticas de préstamos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario y Préstamos"

    # Estilos
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E293B")

    ws.append(["REPORTE DE INVENTARIO Y LIBROS MÁS PRESTADOS"])
    ws.merge_cells("A1:G1")
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([]) # Fila en blanco

    headers = ["ID", "Título", "Autor", "ISBN", "Categoría", "Copias Totales", "Copias Disponibles", "Total Préstamos"]
    ws.append(headers)

    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    libros = Libro.objects.annotate(total_prestamos=Count('prestamos')).order_by('-total_prestamos')

    for libro in libros:
        ws.append([
            libro.id,
            libro.titulo,
            libro.autor,
            libro.isbn,
            str(libro.categoria or "Sin categoría"),
            libro.cantidad_total,
            libro.copias_disponibles,
            libro.total_prestamos
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario_biblioteca.xlsx"'
    return response

def exportar_morosos_excel():
    """Genera un reporte Excel de lectores morosos y préstamos atrasados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lectores Morosos"

    header_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="991B1B")

    ws.append(["REPORTE OFICIAL DE LECTORES MOROSOS Y SANCIONADOS"])
    ws.merge_cells("A1:G1")
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = ["Código Institucional", "Nombre Lector", "Tipo", "Libro Atrasado", "Fecha Límite", "Días Retraso", "Estado Lector"]
    ws.append(headers)

    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    prestamos_atrasados = Prestamo.objects.filter(estado='Atrasado').select_related('lector__user', 'libro')

    for p in prestamos_atrasados:
        nombre = p.lector.user.get_full_name() or p.lector.user.username
        ws.append([
            p.lector.codigo_institucional,
            nombre,
            p.lector.tipo,
            p.libro.titulo,
            p.fecha_limite.strftime('%d/%m/%Y'),
            p.dias_retraso,
            p.lector.estado
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_lectores_morosos.xlsx"'
    return response

def exportar_morosos_csv():
    """Genera un archivo CSV de lectores morosos para fácil interoperabilidad."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="reporte_lectores_morosos.csv"'

    # Añadir BOM para compatibilidad con Excel
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(["Código Institucional", "Nombre Lector", "Tipo", "Libro Atrasado", "Fecha Límite", "Días Retraso", "Estado"])

    prestamos_atrasados = Prestamo.objects.filter(estado='Atrasado').select_related('lector__user', 'libro')
    for p in prestamos_atrasados:
        nombre = p.lector.user.get_full_name() or p.lector.user.username
        writer.writerow([
            p.lector.codigo_institucional,
            nombre,
            p.lector.tipo,
            p.libro.titulo,
            p.fecha_limite.strftime('%d/%m/%Y'),
            p.dias_retraso,
            p.lector.estado
        ])

    return response
